#!/usr/bin/env python3
"""
Scope Database Builder
Scans Y&C preconstruction estimate & clarification letters (.docx)
and extracts scope items into a searchable SQLite database + Excel export.
"""

import os
import re
import sys
import sqlite3
import datetime
from pathlib import Path
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── Config ──────────────────────────────────────────────────────────────────

BASE_DIR = Path(os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-YorkeandCurtis,Inc/01 - PRECONSTRUCTION"
))
SEARCH_DIRS = [
    "01 - PRECONSTRUCTION PROJECTS",
    "02 - BIDDING & RFP PROJECTS",
    "03 - ARCHIVE PROJECTS",
]
DB_PATH = Path(__file__).parent / "scope_database.db"
EXCEL_PATH = Path(__file__).parent / "scope_database.xlsx"

# Patterns to match estimate/clarification files
FILE_PATTERNS = [
    r"estimate.*clarif",
    r"clarif.*estimate",
    r"budget.*clarif",
    r"clarif.*budget",
    r"bid.*clarif",
    r"estimate.*qualification",
    r"proposal",
]
FILE_EXCLUDE = ["template", "~$"]


# ── Division normalization ──────────────────────────────────────────────────

DIVISION_MAP = {
    "1": "Division 1 - General",
    "2": "Division 2 - Sitework",
    "3": "Division 3 - Concrete",
    "4": "Division 4 - Masonry",
    "5": "Division 5 - Metals/Steel",
    "6": "Division 6 - Woods & Plastics",
    "7": "Division 7 - Thermal & Moisture Protection",
    "8": "Division 8 - Doors & Windows",
    "9": "Division 9 - Finishes",
    "10": "Division 10 - Specialties",
    "11": "Division 11 - Equipment",
    "12": "Division 12 - Furnishings",
    "13": "Division 13 - Special Construction",
    "14": "Division 14 - Conveying Systems",
    "15": "Division 15 - Mechanical/Plumbing",
    "16": "Division 16 - Electrical",
}

# ── Helpers ─────────────────────────────────────────────────────────────────

def find_estimate_files():
    """Find all estimate & clarification .docx files."""
    files = []
    for search_dir in SEARCH_DIRS:
        search_path = BASE_DIR / search_dir
        if not search_path.exists():
            continue
        for root, dirs, filenames in os.walk(search_path):
            # Only look in "05 - Quantities" folders
            if "05 - Quantities" not in root and "05 - Quantities & Estimates" not in root:
                # Check if we're inside a quantities folder (subfolder)
                parts = Path(root).parts
                in_quantities = any("05 - Quantities" in p for p in parts)
                if not in_quantities:
                    continue

            for fname in filenames:
                if not fname.endswith(".docx"):
                    continue
                fname_lower = fname.lower()
                # Skip excluded patterns
                if any(exc in fname_lower for exc in FILE_EXCLUDE):
                    continue
                # Match file patterns
                if any(re.search(pat, fname_lower) for pat in FILE_PATTERNS):
                    files.append(os.path.join(root, fname))
    return sorted(set(files))


def extract_project_name_from_path(filepath):
    """Extract project name from the folder structure."""
    parts = Path(filepath).parts
    # Find the part after "01 - PRECONSTRUCTION PROJECTS" or similar
    for i, part in enumerate(parts):
        if part.startswith("01 - PRECON") or part.startswith("02 - BID") or part.startswith("03 - ARCHIVE"):
            # The next part(s) form the project name
            # Skip subcategory folders in archive (like "AP - Apartments")
            remaining = parts[i+1:]
            project_parts = []
            for p in remaining:
                if p.startswith("05 - Quantities"):
                    break
                # Skip archive subcategory folders
                if re.match(r'^[A-Z]{2,3} - ', p) and not any(c.isdigit() for c in p[:5]):
                    continue
                project_parts.append(p)
            if project_parts:
                return " / ".join(project_parts)
    return Path(filepath).stem


def extract_project_category(filepath):
    """Determine if project is Active, Bidding, or Archive."""
    if "01 - PRECONSTRUCTION PROJECTS" in filepath:
        return "Preconstruction"
    elif "02 - BIDDING & RFP PROJECTS" in filepath:
        return "Bidding"
    elif "03 - ARCHIVE PROJECTS" in filepath:
        return "Archive"
    return "Unknown"


def parse_date(text):
    """Try to parse a date from various formats."""
    text = text.strip()
    # Common formats: 04-21-25, 11-20-25, 01-09-26, 7-06-22, 03-16-2022
    patterns = [
        (r'(\d{1,2})-(\d{1,2})-(\d{4})', '%m-%d-%Y'),
        (r'(\d{1,2})-(\d{1,2})-(\d{2})', '%m-%d-%y'),
        (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%m/%d/%Y'),
        (r'(\d{1,2})/(\d{1,2})/(\d{2})', '%m/%d/%y'),
    ]
    for pattern, fmt in patterns:
        m = re.search(pattern, text)
        if m:
            try:
                return datetime.datetime.strptime(m.group(0), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def parse_estimate_amount(text):
    """Extract dollar amount from text."""
    # Match patterns like $1,436,660 or $383,396
    m = re.search(r'\$[\d,]+(?:\.\d{2})?', text)
    if m:
        amount_str = m.group(0).replace('$', '').replace(',', '')
        try:
            return float(amount_str)
        except ValueError:
            pass
    return None


def classify_scope_item(text):
    """Classify a scope item as includes, excludes, assumes, allowance, recommends, or note."""
    text_lower = text.lower().strip()
    if re.search(r'\b(does not include|excludes?|not included|no\b.*included)', text_lower):
        return "excludes"
    elif re.search(r'\b(includes?|included)\b', text_lower):
        if re.search(r'\ballowance\b', text_lower):
            return "allowance"
        return "includes"
    elif re.search(r'\bassumes?\b', text_lower):
        return "assumes"
    elif re.search(r'\brecommend\b', text_lower):
        return "recommends"
    elif re.search(r'\ballowance\b', text_lower):
        return "allowance"
    return "note"


def normalize_division(text):
    """Normalize division header text to standard form."""
    text = text.strip()
    # Extract division number
    m = re.search(r'division\s+(\d+)', text, re.IGNORECASE)
    if m:
        div_num = m.group(1)
        return DIVISION_MAP.get(div_num, f"Division {div_num}")
    return text


def parse_docx(filepath):
    """Parse a single estimate & clarifications docx file."""
    try:
        doc = Document(filepath)
    except Exception as e:
        return None, f"Error opening: {e}"

    paragraphs = [(p.text.strip(), p.style.name if p.style else "Normal") for p in doc.paragraphs]

    result = {
        "date": None,
        "recipient": None,
        "subject": None,
        "estimate_amount": None,
        "estimate_type": None,  # Preliminary, Budget, Final, etc.
        "scope_items": [],       # list of (division, item_text, item_type)
        "standard_exclusions": [],
    }

    current_division = None
    in_standard_exclusions = False
    found_clarifications_header = False

    for i, (text, style) in enumerate(paragraphs):
        if not text:
            continue

        # Date - usually first non-empty line
        if result["date"] is None and i < 5:
            d = parse_date(text)
            if d:
                result["date"] = d
                continue

        # Subject line (Re:)
        if text.lower().startswith("re:") and result["subject"] is None:
            result["subject"] = text[3:].strip()
            # Extract estimate type
            subj_lower = text.lower()
            if "preliminary" in subj_lower or "prelim" in subj_lower:
                result["estimate_type"] = "Preliminary"
            elif "rom" in subj_lower:
                result["estimate_type"] = "ROM"
            elif "budget" in subj_lower:
                result["estimate_type"] = "Budget"
            elif "bid" in subj_lower:
                result["estimate_type"] = "Bid"
            else:
                result["estimate_type"] = "Estimate"
            continue

        # Estimate amount - look for dollar amount in Title/Normal style near top
        if result["estimate_amount"] is None and i < 30:
            amt = parse_estimate_amount(text)
            if amt and amt > 1000:  # Filter out small numbers
                result["estimate_amount"] = amt
                # Also check for estimate type if not found yet
                if result["estimate_type"] is None:
                    text_lower = text.lower()
                    if "preliminary" in text_lower:
                        result["estimate_type"] = "Preliminary"
                    elif "rom" in text_lower:
                        result["estimate_type"] = "ROM"
                    elif "budget" in text_lower:
                        result["estimate_type"] = "Budget"

        # Dear line - extract recipient
        if text.lower().startswith("dear") and result["recipient"] is None:
            result["recipient"] = text
            continue

        # Standard exclusions section
        if re.search(r'standard\s+exclusion', text, re.IGNORECASE):
            in_standard_exclusions = True
            current_division = None
            continue

        # Division headers
        div_match = re.match(r'Division\s+\d+', text, re.IGNORECASE)
        if div_match:
            current_division = normalize_division(text)
            in_standard_exclusions = False
            continue

        # Clarifications header
        if re.search(r'specific\s+qualifications?\s*&?\s*clarifications?', text, re.IGNORECASE):
            found_clarifications_header = True
            continue

        # Scope items (list paragraphs or bullets under a division)
        if style == "List Paragraph" or (text.startswith("•") or text.startswith("-") or text.startswith("–")):
            clean_text = text.lstrip("•-–· \t")
            if not clean_text or len(clean_text) < 5:
                continue

            if in_standard_exclusions:
                result["standard_exclusions"].append(clean_text)
            elif current_division:
                item_type = classify_scope_item(clean_text)
                result["scope_items"].append((current_division, clean_text, item_type))
            elif found_clarifications_header:
                # Items before first division header but after clarifications
                item_type = classify_scope_item(clean_text)
                result["scope_items"].append(("General Notes", clean_text, item_type))

    # Try to get date from filename if not found in document
    if result["date"] is None:
        d = parse_date(os.path.basename(filepath))
        if d:
            result["date"] = d

    return result, None


# ── Database ────────────────────────────────────────────────────────────────

def create_database():
    """Create the SQLite database schema."""
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()

    c.execute("DROP TABLE IF EXISTS scope_items")
    c.execute("DROP TABLE IF EXISTS documents")
    c.execute("DROP TABLE IF EXISTS standard_exclusions")

    c.execute("""
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT,
            project_category TEXT,
            file_path TEXT,
            file_name TEXT,
            document_date TEXT,
            estimate_amount REAL,
            estimate_type TEXT,
            subject_line TEXT,
            recipient TEXT
        )
    """)

    c.execute("""
        CREATE TABLE scope_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER,
            project_name TEXT,
            project_category TEXT,
            document_date TEXT,
            division TEXT,
            scope_text TEXT,
            scope_type TEXT,
            FOREIGN KEY (document_id) REFERENCES documents(id)
        )
    """)

    c.execute("""
        CREATE TABLE standard_exclusions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER,
            exclusion_text TEXT,
            FOREIGN KEY (document_id) REFERENCES documents(id)
        )
    """)

    # Full-text search index on scope items
    c.execute("""
        CREATE VIRTUAL TABLE IF NOT EXISTS scope_items_fts USING fts5(
            project_name,
            division,
            scope_text,
            scope_type,
            content='scope_items',
            content_rowid='id'
        )
    """)

    conn.commit()
    return conn


def populate_database(conn, files):
    """Process all files and populate the database."""
    c = conn.cursor()
    stats = {"files": 0, "items": 0, "errors": 0, "skipped": 0}

    for filepath in files:
        result, error = parse_docx(filepath)
        if error:
            print(f"  ERROR: {os.path.basename(filepath)}: {error}")
            stats["errors"] += 1
            continue

        if not result["scope_items"]:
            stats["skipped"] += 1
            continue

        project_name = extract_project_name_from_path(filepath)
        project_category = extract_project_category(filepath)

        c.execute("""
            INSERT INTO documents (project_name, project_category, file_path, file_name,
                                   document_date, estimate_amount, estimate_type,
                                   subject_line, recipient)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            project_name, project_category, filepath, os.path.basename(filepath),
            result["date"], result["estimate_amount"], result["estimate_type"],
            result["subject"], result["recipient"],
        ))
        doc_id = c.lastrowid

        for division, text, item_type in result["scope_items"]:
            c.execute("""
                INSERT INTO scope_items (document_id, project_name, project_category,
                                         document_date, division, scope_text, scope_type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (doc_id, project_name, project_category, result["date"],
                  division, text, item_type))

        for excl in result["standard_exclusions"]:
            c.execute("""
                INSERT INTO standard_exclusions (document_id, exclusion_text)
                VALUES (?, ?)
            """, (doc_id, excl))

        stats["files"] += 1
        stats["items"] += len(result["scope_items"])

    # Populate FTS index
    c.execute("""
        INSERT INTO scope_items_fts (rowid, project_name, division, scope_text, scope_type)
        SELECT id, project_name, division, scope_text, scope_type FROM scope_items
    """)

    conn.commit()
    return stats


# ── Excel Export ────────────────────────────────────────────────────────────

def export_excel(conn):
    """Export the database to a formatted Excel workbook."""
    wb = Workbook()
    c = conn.cursor()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    type_colors = {
        "includes": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "excludes": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "assumes": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "allowance": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "recommends": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
        "note": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }

    # ── Sheet 1: All Scope Items ──
    ws = wb.active
    ws.title = "Scope Items"
    headers = ["Project", "Category", "Date", "Division", "Scope Item", "Type", "Estimate Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    c.execute("""
        SELECT si.project_name, si.project_category, si.document_date, si.division,
               si.scope_text, si.scope_type, d.estimate_amount
        FROM scope_items si
        JOIN documents d ON si.document_id = d.id
        ORDER BY si.project_name, si.document_date, si.division, si.id
    """)

    for row_idx, row in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 7 and val:  # Estimate amount
                cell.number_format = '$#,##0'
        # Color-code by scope type
        scope_type = row[5]
        if scope_type in type_colors:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = type_colors[scope_type]

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16

    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.freeze_panes = "A2"

    # ── Sheet 2: Documents Summary ──
    ws2 = wb.create_sheet("Documents")
    doc_headers = ["Project", "Category", "Date", "Estimate Type", "Estimate Amount",
                   "# Scope Items", "File Name"]
    for col, h in enumerate(doc_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    c.execute("""
        SELECT d.project_name, d.project_category, d.document_date, d.estimate_type,
               d.estimate_amount, COUNT(si.id), d.file_name
        FROM documents d
        LEFT JOIN scope_items si ON d.id = si.document_id
        GROUP BY d.id
        ORDER BY d.project_name, d.document_date
    """)

    for row_idx, row in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 5 and val:
                cell.number_format = '$#,##0'

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 50

    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Division Summary (pivot-style) ──
    ws3 = wb.create_sheet("By Division")
    div_headers = ["Division", "Total Items", "Includes", "Excludes", "Assumes",
                   "Allowances", "Recommends", "Notes", "# Projects"]
    for col, h in enumerate(div_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    c.execute("""
        SELECT division,
               COUNT(*),
               SUM(CASE WHEN scope_type='includes' THEN 1 ELSE 0 END),
               SUM(CASE WHEN scope_type='excludes' THEN 1 ELSE 0 END),
               SUM(CASE WHEN scope_type='assumes' THEN 1 ELSE 0 END),
               SUM(CASE WHEN scope_type='allowance' THEN 1 ELSE 0 END),
               SUM(CASE WHEN scope_type='recommends' THEN 1 ELSE 0 END),
               SUM(CASE WHEN scope_type='note' THEN 1 ELSE 0 END),
               COUNT(DISTINCT project_name)
        FROM scope_items
        GROUP BY division
        ORDER BY division
    """)

    for row_idx, row in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws3.column_dimensions['A'].width = 40
    for col in 'BCDEFGHI':
        ws3.column_dimensions[col].width = 14

    ws3.freeze_panes = "A2"

    wb.save(str(EXCEL_PATH))


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Y&C Scope Database Builder")
    print("=" * 60)

    print(f"\nScanning: {BASE_DIR}")
    files = find_estimate_files()
    print(f"Found {len(files)} estimate & clarification documents")

    if not files:
        print("No files found!")
        sys.exit(1)

    print(f"\nCreating database: {DB_PATH}")
    conn = create_database()

    print("Processing documents...")
    stats = populate_database(conn, files)

    print(f"\nResults:")
    print(f"  Documents processed: {stats['files']}")
    print(f"  Scope items extracted: {stats['items']}")
    print(f"  Files with errors: {stats['errors']}")
    print(f"  Files skipped (no items): {stats['skipped']}")

    # Quick stats
    c = conn.cursor()
    c.execute("SELECT COUNT(DISTINCT project_name) FROM documents")
    print(f"  Unique projects: {c.fetchone()[0]}")
    c.execute("SELECT COUNT(DISTINCT division) FROM scope_items")
    print(f"  Divisions covered: {c.fetchone()[0]}")

    print(f"\nExporting Excel: {EXCEL_PATH}")
    export_excel(conn)

    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
